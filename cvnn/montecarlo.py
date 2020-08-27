import logging
import cvnn
import cvnn.layers as layers
import cvnn.dataset as dp
from cvnn.cvnn_model import CvnnModel
from cvnn.data_analysis import MonteCarloAnalyzer
from cvnn.layers import Dense
from cvnn.utils import transform_to_real, randomize
from tensorflow.keras.losses import categorical_crossentropy
import pandas as pd
import copy
import sys
from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table
import os
from tqdm import tqdm
import numpy as np
from pdb import set_trace
from time import sleep

logger = logging.getLogger(cvnn.__name__)


class MonteCarlo:

    def __init__(self):
        self.models = []
        self.pandas_full_data = pd.DataFrame()
        self.confusion_matrix = []
        self.monte_carlo_analyzer = MonteCarloAnalyzer()  # All at None

    def add_model(self, model):
        self.models.append(model)

    def run(self, x, y, data_summary='', polar=False, do_conf_mat=True, validation_split=0.2,
            iterations=100, learning_rate=0.01, epochs=10, batch_size=100,
            shuffle=False, debug=False, display_freq=1, checkpoints=False):
        x, y = randomize(x, y)
        # Reset data frame
        self.pandas_full_data = pd.DataFrame()
        if do_conf_mat:
            for i in range(len(self.models)):
                self.confusion_matrix.append({"name": "model_name", "matrix": pd.DataFrame()})
        self.save_summary_of_run(self._run_summary(iterations, learning_rate, epochs, batch_size, shuffle),
                                 data_summary)
        if not debug:
            pbar = tqdm(total=iterations)
        for it in range(iterations):
            if debug:
                logger.info("Iteration {}/{}".format(it + 1, iterations))
            else:
                pbar.update()
            if shuffle:  # shuffle all data at each iteration
                x, y = randomize(x, y)
            for i, model in enumerate(self.models):
                if model.is_complex():
                    x_fit = x
                else:
                    x_fit = transform_to_real(x, polar=polar)
                test_model = copy.deepcopy(model)
                test_model.fit(x_fit, y, validation_split=validation_split,
                               learning_rate=learning_rate, epochs=epochs, batch_size=batch_size,
                               verbose=debug, fast_mode=True, save_txt_fit_summary=False, display_freq=display_freq,
                               save_csv_history=True)
                self.pandas_full_data = pd.concat([self.pandas_full_data,
                                                   test_model.plotter.get_full_pandas_dataframe()], sort=False)
                if do_conf_mat:
                    dataset = dp.Dataset(x_fit, y, ratio=1-validation_split)
                    self.confusion_matrix[i]["name"] = test_model.name
                    self.confusion_matrix[i]["matrix"] = pd.concat((self.confusion_matrix[i]["matrix"],
                                                                    test_model.get_confusion_matrix(dataset.x_test,
                                                                                                    dataset.y_test)))
            if checkpoints:
                # Save checkpoint in case Monte Carlo stops in the middle
                self.pandas_full_data.to_csv(self.monte_carlo_analyzer.path / "run_data.csv", index=False)
        if not debug:
            pbar.close()
        self.pandas_full_data = self.pandas_full_data.reset_index(drop=True)
        conf_mat = None
        if do_conf_mat:
            conf_mat = self.confusion_matrix
        self.monte_carlo_analyzer.set_df(self.pandas_full_data, conf_mat)

    @staticmethod
    def _run_summary(iterations, learning_rate, epochs, batch_size, shuffle):
        ret_str = "Monte Carlo run\n"
        ret_str += "\tIterations: {}\n".format(iterations)
        ret_str += "\tepochs: {}\n".format(epochs)
        ret_str += "\tbatch_size: {}\n".format(batch_size)
        ret_str += "\tLearning Rate: {}\n".format(learning_rate)
        if shuffle:
            ret_str += "\tShuffle data at each iteration\n"
        else:
            ret_str += "\tData is not shuffled at each iteration\n"
        return ret_str

    def save_summary_of_run(self, run_summary, data_summary):
        with open(str(self.monte_carlo_analyzer.path / "run_summary.txt"), "w") as file:
            file.write(run_summary)
            file.write(data_summary)
            file.write("Models:\n")
            for model in self.models:
                file.write(model.summary())


class RealVsComplex(MonteCarlo):

    def __init__(self, complex_model, capacity_equivalent=True, equiv_technique='ratio'):
        super().__init__()
        # add models
        self.add_model(complex_model)
        self.add_model(complex_model.get_real_equivalent(capacity_equivalent=capacity_equivalent,
                                                         equiv_technique=equiv_technique, name="real_network"))


# ====================================
#     Excel logging
# ====================================
def run_montecarlo(models, dataset, open_dataset=None, iterations=500,
                   epochs=150, batch_size=100, display_freq=1, learning_rate=0.01,
                   debug=False, polar=False, do_all=True):
    if open_dataset:
        dataset = dp.OpenDataset(open_dataset)  # Warning, open_dataset overwrites dataset

    # Monte Carlo
    monte_carlo = MonteCarlo()
    for model in models:
        monte_carlo.add_model(model)
    if not open_dataset:
        dataset.save_data(monte_carlo.monte_carlo_analyzer.path)
    monte_carlo.run(dataset.x, dataset.y, iterations=iterations, learning_rate=learning_rate,
                    epochs=epochs, batch_size=batch_size, display_freq=display_freq,
                    shuffle=False, debug=debug, data_summary=dataset.summary(), polar=polar)
    if do_all:
        monte_carlo.monte_carlo_analyzer.do_all()

    # Save data to remember later what I did.
    _save_montecarlo_log(path=str(monte_carlo.monte_carlo_analyzer.path),
                         models_names=[str(model.name) for model in models],
                         dataset_name=dataset.dataset_name,
                         num_classes=str(dataset.y.shape[1]),
                         polar_mode='Yes' if polar == 'Apple' else 'No',
                         learning_rate=learning_rate,
                         dataset_size=str(dataset.x.shape[0]),
                         features_size=str(dataset.x.shape[1]), epochs=epochs, batch_size=batch_size
                         )
    return str(monte_carlo.monte_carlo_analyzer.path / "run_data.csv")


def run_gaussian_dataset_montecarlo(iterations=1000, m=10000, n=128, param_list=None,
                                    epochs=150, batch_size=100, display_freq=1, learning_rate=0.01,
                                    shape_raw=None, activation='cart_relu', debug=False, polar=False, do_all=True,
                                    dropout=None):
    # Get parameters
    if param_list is None:
        param_list = [
            [0.5, 1, 1],
            [-0.5, 1, 1]
        ]
    dataset = dp.CorrelatedGaussianCoeffCorrel(m, n, param_list, debug=False)
    mlp_run_real_comparison_montecarlo(dataset, None, iterations, epochs, batch_size, display_freq, learning_rate,
                                       shape_raw, activation, debug, polar, do_all, dropout=dropout)


def mlp_run_real_comparison_montecarlo(dataset, open_dataset=None, iterations=1000,
                                       epochs=150, batch_size=100, display_freq=1, learning_rate=0.01,
                                       shape_raw=None, activation='cart_relu',
                                       debug=False, polar=False, do_all=True, dropout=0.5):
    if shape_raw is None:
        shape_raw = [64]
    if open_dataset:
        dataset = dp.OpenDataset(open_dataset)  # Warning, open_dataset overwrites dataset
    # Create complex network
    input_size = dataset.x.shape[1]  # Size of input
    output_size = dataset.y.shape[1]  # Size of output
    if not len(shape_raw) > 0:
        logger.error("Shape raw was empty")
        sys.exit(-1)
    layers.ComplexLayer.last_layer_output_dtype = None
    layers.ComplexLayer.last_layer_output_size = None
    if len(shape_raw) == 0:
        logger.warning("No hidden layers are used. activation and dropout will be ignored")
        shape = [
            Dense(input_size=input_size, output_size=output_size, activation='softmax_real',
                  input_dtype=np.complex64, dropout=None)
        ]
    else:   # len(shape_raw) > 0:
        shape = [Dense(input_size=input_size, output_size=shape_raw[0], activation=activation,
                       input_dtype=np.complex64, dropout=dropout)]
        for i in range(1, len(shape_raw)):
            shape.append(Dense(output_size=shape_raw[i], activation=activation, dropout=dropout))
        shape.append(Dense(output_size=output_size, activation='softmax_real', dropout=None))

    complex_network = CvnnModel(name="complex_network", shape=shape, loss_fun=categorical_crossentropy,
                                verbose=False, tensorboard=False)

    # Monte Carlo
    monte_carlo = RealVsComplex(complex_network, capacity_equivalent=False)
    if not open_dataset:
        dataset.save_data(monte_carlo.monte_carlo_analyzer.path)
    sleep(1)  # I have error if not because not enough time passed since creation of models to be in diff folders
    monte_carlo.run(dataset.x, dataset.y, iterations=iterations, learning_rate=learning_rate,
                    epochs=epochs, batch_size=batch_size, display_freq=display_freq,
                    shuffle=False, debug=debug, data_summary=dataset.summary(), polar=polar)
    if do_all:
        monte_carlo.monte_carlo_analyzer.do_all()

    # Save data to remember later what I did.
    max_epoch = monte_carlo.pandas_full_data['epoch'].max()
    epoch_filter = monte_carlo.pandas_full_data['epoch'] == max_epoch
    complex_filter = monte_carlo.pandas_full_data['network'] == "complex network"
    real_filter = monte_carlo.pandas_full_data['network'] == "real network"
    complex_last_epochs = monte_carlo.pandas_full_data[epoch_filter & complex_filter]
    real_last_epochs = monte_carlo.pandas_full_data[epoch_filter & real_filter]
    complex_median = complex_last_epochs['test accuracy'].median()
    real_median = real_last_epochs['test accuracy'].median()
    _save_rvnn_vs_cvnn_montecarlo_log(path=str(monte_carlo.monte_carlo_analyzer.path),
                                      dataset_name=dataset.dataset_name,
                                      hl=str(len(shape_raw)), shape=str(shape_raw),
                                      dropout=str(dropout), num_classes=str(dataset.y.shape[1]),
                                      polar_mode='Yes' if polar else 'No',
                                      learning_rate=learning_rate, activation=activation,
                                      dataset_size=str(dataset.x.shape[0]),
                                      feature_size=str(dataset.x.shape[1]), epochs=epochs, batch_size=batch_size,
                                      winner='CVNN' if complex_median > real_median else 'RVNN',
                                      complex_median=complex_median, real_median=real_median,
                                      complex_iqr=complex_last_epochs['test accuracy'].quantile(.75)
                                                  - complex_last_epochs['test accuracy'].quantile(.25),
                                      real_iqr=real_last_epochs['test accuracy'].quantile(.75)
                                               - real_last_epochs['test accuracy'].quantile(.25)
                                      )
    return str(monte_carlo.monte_carlo_analyzer.path / "run_data.csv")


# ====================================
#     Excel logging
# ====================================
def _create_excel_file(fieldnames, row_data, filename=None, percentage_cols=None):
    if filename is None:
        filename = './log/montecarlo_summary.xlsx'
    file_exists = os.path.isfile(filename)
    if file_exists:
        wb = load_workbook(filename)
        ws = wb.worksheets[0]
    else:
        wb = Workbook()
        ws = wb.worksheets[0]
        ws.append(fieldnames)
    ws.append(row_data)
    # TODO: What if len(row_data) is longer than the dictionary? It corresponds with excel's column names?
    tab = Table(displayName="Table1", ref="A1:" + str(chr(64 + len(row_data))) + str(ws.max_row))
    if percentage_cols is not None:
        for col in percentage_cols:
            ws[col + str(ws.max_row)].number_format = '0.00%'
    ws.add_table(tab)
    wb.save(filename)


def _save_rvnn_vs_cvnn_montecarlo_log(path, dataset_name, hl, shape, dropout, num_classes, polar_mode, learning_rate,
                                      activation,
                                      dataset_size, feature_size, epochs, batch_size, winner,
                                      complex_median, real_median, complex_iqr, real_iqr, comments='', filename=None):
    fieldnames = ['dataset', '# Classes', "Dataset Size", 'Feature Size', "Polar Mode",
                  'HL', 'Shape', 'Dropout', "Learning Rate", "Activation Function", 'epochs', 'batch size',
                  "Winner", "CVNN median", "RVNN median", 'CVNN IQR', 'RVNN IQR',
                  'path', "cvnn version", "Comments"
                  ]
    row_data = [dataset_name, num_classes, dataset_size, feature_size, polar_mode,  # Dataset information
                hl, shape, dropout, learning_rate, activation, epochs, batch_size,  # Model information
                winner, complex_median, real_median, complex_iqr, real_iqr,  # Preliminary results
                path, cvnn.__version__, comments  # Library information
                ]
    percentage_cols = ['N', 'O', 'P', 'Q']
    _create_excel_file(fieldnames, row_data, filename, percentage_cols=percentage_cols)


def _save_montecarlo_log(path, dataset_name, models_names, num_classes, polar_mode, learning_rate, dataset_size,
                         features_size, epochs, batch_size, filename=None):
    fieldnames = [
        'dataset', '# Classes', "Dataset Size", 'Feature Size',  # Dataset information
        'models', 'epochs', 'batch size', "Polar Mode", "Learning Rate",  # Models information
        'path', "cvnn version"  # Library information
    ]
    row_data = [
        dataset_name, num_classes, dataset_size, features_size,
        '-'.join(models_names), epochs, batch_size, polar_mode, learning_rate,
        path, cvnn.__version__
    ]
    _create_excel_file(fieldnames, row_data, filename)


if __name__ == "__main__":
    # Base case with one hidden layer size 64 and dropout 0.5
    run_gaussian_dataset_montecarlo(iterations=10, dropout=None)
